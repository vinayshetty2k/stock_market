from packages.nse import Nse
import pprint
import sys
import pandas as pd
import openpyxl
import os


def save_excel(df):
    writer = pd.ExcelWriter(wb_name, engine='openpyxl')
    if (os.path.exists(wb_name)):
        workbook = openpyxl.load_workbook(wb_name)
        writer.book = workbook
    df.to_excel(writer, sheet_name=sheetname,index=False)
    worksheet = writer.sheets[sheetname]
    for col, column_cells  in enumerate(worksheet.columns):
        cell = worksheet.cell(row=1, column=col+1)
        column_width = len(as_text(cell.value))+2
        if column_width > 50:    column_width = 50
        if column_width < 10:    column_width = 10
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = column_width
    writer.save
    writer.close()


def as_text(value):
    if value is None:
        return ""
    return str(value)


functionDict = {"active_stocks": "get_active_monthly", "top_gainers": "get_top_gainers", \
                "top_losers": "get_top_losers", "preopen_all": "get_preopen_all", \
                "preopen_futures": "get_preopen_fno", "preopen_nifty": "get_preopen_nifty", \
                "preopen_niftybank": "get_preopen_niftybank", "adv_dec": "get_advances_declines", \
                "yearHigh": "get_year_high", "yearLow": "get_year_low", "json_data": "get_json_data", "csv_data": "get_csv_data"}

function = sys.argv[1].strip()
nse = Nse()
wb_name = os.path.join(os.path.dirname(sys.argv[0]), 'Miscellaneous.xlsx')

if function == "json_data" or function == "csv_data":
    url = sys.argv[2].replace("replacAmp", "&")
    nse_data = getattr(nse, functionDict[function])(url)
    sheetname = sys.argv[3]
    if function == "csv_data": wb_name = os.path.join(os.path.dirname(sys.argv[0]), 'Future_Historical_Data.xlsx')
else:
    method_to_call = getattr(nse, functionDict[function])
    nse_data = method_to_call()
    sheetname = function

if function == "csv_data":
    df = pd.read_csv(nse_data)
else:
    df = pd.DataFrame(nse_data)
save_excel(df)
